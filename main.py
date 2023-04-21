import re
import PyPDF2
import tabula
import pandas as pd
import numpy as np
import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from colour import Color
import argparse

def format_excel_file(file_path):
    # Load the workbook
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Rename the first cell in column K to 'DEPT' and the first cell in column L to 'Q.C.'
    ws["K1"].value = 'DEPT'
    ws["L1"].value = 'Q.C.'

    # Calculate text color based on background color
    def get_text_color(bg_color_hex):
        bg_color = Color(bg_color_hex)
        luma = 0.299 * bg_color.red * 255 + 0.587 * bg_color.green * 255 + 0.114 * bg_color.blue * 255
        return "FFFFFF" if luma < 128 else "000000"

    # Apply bold font, Helvetica, and the desired fill color to header row
    highlight_color = PatternFill(start_color="B02522", end_color="B02522", fill_type="solid")
    for row_cells in ws["A1:L1"]:
        for cell in row_cells:
            cell.font = Font(bold=True, name="Helvetica", color=get_text_color("#B02522"))
            cell.fill = highlight_color

    for row in ws.iter_rows(min_row=2):
        if row[0].value:  # Check if the first cell in the row (column A) is not empty
            row[0].font = Font(size=14, bold=True, name="Helvetica", color=get_text_color("#B02522"))
            ws.row_dimensions[row[0].row].height = 20  # Set the height of the highlighted row
            # Merge cells of the highlighted row
            ws.merge_cells(start_row=row[0].row, start_column=1, end_row=row[0].row, end_column=12)

            for i, cell in enumerate(row[:12]):  # Iterate through cells in the row from column A to K
                cell.fill = highlight_color
                if i > 0:
                    cell.font = Font(name="Helvetica", color=get_text_color("#B02522"))

                # Apply thick border to the top and bottom edges of all cells
                top_border = Side(style='thick')
                bottom_border = Side(style='thick')

                # Apply thick border to the left edge of the first cell (column A)
                left_border = Side(style='thick') if i == 0 else None

                # Apply thick border to the right edge of the last cell (column L)
                right_border = Side(style='thick') if i == 11 else None

                cell.border = Border(left=left_border, right=right_border, top=top_border, bottom=bottom_border)
        else:
            for cell in row:  # Iterate through cells in the row
                thin_border = Side(style='thin')
                cell.border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
                if cell.fill.start_color.index == "FFFFFFFF":
                    cell.border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)

    # Set the font to Helvetica for the rest of the sheet
    for row in ws.iter_rows(min_row=2):
        for cell in row[11:]:
            cell.font = Font(name="Helvetica")

    # Round numbers in columns D and E to two decimal places after converting text to float
    for row in ws.iter_rows(min_row=2):
        if row[3].value and isinstance(row[3].value, str) and row[3].value.replace('.', '', 1).isdigit():
            row[3].value = "{:.2f}".format(round(float(row[3].value), 2))  # Column D
        if row[4].value and isinstance(row[4].value, str) and row[4].value.replace('.', '', 1).isdigit():
            row[4].value = "{:.2f}".format(round(float(row[4].value), 2))  # Column E

    # Center the text in columns B and I
    for row in ws.iter_rows():
        row[1].alignment = Alignment(horizontal='center')  # Column B
        row[8].alignment = Alignment(horizontal='center')  # Column I
        # Set column widths in number of characters
        column_widths = {
            'A': 6,
            'B': 6,
            'C': 36,
            'D': 10,
            'E': 10,
            'F': 10,
            'G': 10,
            'H': 10,
            'I': 6,
            'J': 33,
            'K': 6,
            'L': 6,
            'M': 20,
            'N': 20,
        }

    for col_letter, width_chars in column_widths.items():
        ws.column_dimensions[col_letter].width = width_chars

    # Freeze the top row
    ws.freeze_panes = "A2"

    # Apply conditional formatting rules
    # Column K: Turn row green if cell contains "P" or "p"
    green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    rule = FormulaRule(formula=['OR(LOWER($K1)="p"'], stopIfTrue=True, fill=green_fill)
    ws.conditional_formatting.add('A1:K1048575', rule)

    # Column K: Turn row yellow if cell contains "F" or "f"
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    rule = FormulaRule(formula=['LOWER($K1)="f"'], stopIfTrue=True, fill=yellow_fill)
    ws.conditional_formatting.add('A1:K1048575', rule)

    # Column L: Turn cell red if the number is negative
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    rule = CellIsRule(operator="lessThan", formula=['0'], fill=red_fill)
    ws.conditional_formatting.add('L2:L1048576', rule)

    # Column L: Turn cell orange if the number is positive
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    rule = CellIsRule(operator="greaterThan", formula=['0'], fill=orange_fill)
    ws.conditional_formatting.add('L2:L1048576', rule)
    ws.column_dimensions['J'].hidden = True
    wb.active.title = os.path.splitext(os.path.basename(file_path))[0]

    # Set sheet title to the filename
    ws.title = os.path.splitext(os.path.basename(file_path))[0]

    # Add file name to cell M1
    ws['M1'] = os.path.basename(file_path)
    # Merge cells M1 and N1
    ws.merge_cells('M1:N1')

    # Add "Inspector Name:" to cell M2 and "Date" to cell M3
    ws['M2'] = "Inspector Name:"
    ws['M3'] = "Date"

    heavy_border = Side(style='thick')
    thin_border = Side(style='thin')

    # Borders for cells M1 and N1
    ws['M1'].border = Border(left=heavy_border, top=heavy_border, right=thin_border)
    ws['N1'].border = Border(top=heavy_border, right=heavy_border, left=thin_border)

    for row in ws['M2:N3']:
        for cell in row:
            left_border = heavy_border if cell.column == 'M' else thin_border
            right_border = None
            top_border = thin_border
            bottom_border = heavy_border if cell.row == 3 else thin_border

            cell.border = Border(left=left_border, right=right_border, top=top_border, bottom=bottom_border)

    # Add missing thick borders on the left side of O2 and O3
    ws['O2'].border = Border(left=heavy_border)
    ws['O3'].border = Border(left=heavy_border)

    # Get the file name without the extension and everything after the dash
    file_name = os.path.splitext(os.path.basename(file_path))[0]
    file_name = file_name.split('-')[0].strip()

    # Add the formatted file name to the merged cells M1:N1
    ws['M1'] = file_name
    ws['M1'].font = Font(bold=True)
    ws['M1'].alignment = Alignment(horizontal="center")

    # Set "Inspector Name:" and "Date" to bold
    ws['M2'].font = Font(bold=True)
    ws['M3'].font = Font(bold=True)

    wb.save(file_path)


def mm_to_points(mm):
    return mm * 2.83465


def extract_tables_tabula(doc, area_mm, column_positions_mm):
    area_points = [mm_to_points(x) for x in area_mm]
    column_positions_points = [mm_to_points(x) for x in column_positions_mm]
    df_list = tabula.read_pdf(doc, pages='all', multiple_tables=True, stream=True, guess=True, area=area_points,
                              columns=column_positions_points)
    df = pd.concat(df_list)
    print("Extracting")
    return df


def clean_dataframe(df):
    # Rename columns
    df.columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']

    # Combine columns A and B if "PSI I" is found in column A
    df['A'] = df.apply(
        lambda row: str(row['A']).strip() + str(row['B']).strip() if "PSI I" in str(row['A']) else row['A'], axis=1)

    df.loc[df['A'].str.contains("PSI I", na=False), 'B'] = ''

    # Replace row contents with empty strings if "QTY" in A and "Material" in G
    df.loc[(df['A'].str.contains("QTY", na=False)) & (df['G'].str.contains("Material", na=False))] = ''

    # Look for d"Requiring Bottom" in B and make the entire row empty when it's found
    df.loc[df['B'].str.contains("Requiring Bottom", na=False)] = ''

    # Shift columns F, G, H, and I one column to the left
    df[['E', 'F', 'G', 'H']] = df[['F', 'G', 'H', 'I']]
    df['I'] = np.nan  # Set column I to NaN

    # ... other cleaning steps

    # Insert a new column between A and B
    df.insert(1, 'QTY', np.nan)

    # Insert a new column between E and F
    df.insert(6, 'New_Column2', np.nan)

    # Insert a new column between F and G
    df.insert(8, 'New_Column3', np.nan)

    # Replace the contents of column D with the contents of column E
    df['D'] = df['E']

    # Move the cell in column A to column QTY if it does not contain 'PSI'
    df['QTY'] = df.apply(lambda row: row['A'] if 'PSI' not in str(row['A']) else np.nan, axis=1)
    df.loc[df['QTY'].notna(), 'A'] = np.nan

    new_column_names = ['PSI ID', 'QTY', 'DESCRIPTION', 'WIDTH', 'LENGTH', 'CORE', 'INTERIOR', 'EXTERIOR', 'BLANK',
                        'CODE', 'SEQ#', 'BLANK2']
    df.columns = new_column_names[:len(df.columns)]  # Assign the new column names up to the number of existing columns

    with open("core_materials.txt", "r") as file:
        search_strings = [line.strip() for line in file.readlines()]

    # Define a function to search for a matching string in the list and split the given string into three parts
    def split_string(s, search_strings):
        s = str(s)  # Convert the input to a string
        for search_string in search_strings:
            if search_string in s:
                first_part = s.split(search_string)[0].strip()
                middle_part = search_string
                last_part = s.split(search_string)[1].strip()
                return first_part, middle_part, last_part
        return s, None, None  # Return the original string and None for the other parts when there's no match

        # Apply the function to the "EXTERIOR" column

    split_results = df['EXTERIOR'].apply(lambda x: pd.Series(split_string(x, search_strings)))

    # Update only rows with a match
    mask = (split_results[0] != df['EXTERIOR'])  # Identify rows with a match
    df.loc[mask, 'INTERIOR'] = split_results.loc[mask, 0]
    df.loc[mask, 'CORE'] = split_results.loc[mask, 1]
    df.loc[mask, 'EXTERIOR'] = split_results.loc[mask, 2]

    # ...rest of your code
    df = df.applymap(lambda x: '' if pd.isnull(x) or x == 'nan' else x)
    df.loc[df['QTY'] == '', 'CODE'] = ''
    df.loc[df['DESCRIPTION'].str.contains("Part Description", na=False)] = ''
    df = df[~((df['WIDTH'] == '') & (df['LENGTH'] == '') & (df['CORE'] == '') & (df['PSI ID'] == '') & (
        ~df['DESCRIPTION'].str.contains("Parts Per Reports: ", na=False)))]
    df['DESCRIPTION'] = df.apply(
        lambda row: str(row['QTY']).strip() + ' ' + str(row['DESCRIPTION']).strip() if "Parts Per Reports:" in str(
            row['DESCRIPTION']) and "Total" in str(row['QTY']) else row['DESCRIPTION'], axis=1)
    df.loc[(df['DESCRIPTION'].str.contains("Parts Per Reports:", na=False)) & (
        df['QTY'].str.contains("Total", na=False)), 'QTY'] = ''

    # Convert the QTY column to numeric values, ignoring non-numeric values
    df['QTY'] = pd.to_numeric(df['QTY'], errors='coerce')

    # Calculate the sum of the QTY column, skipping NaN values
    total_qty = df['QTY'].sum(skipna=True)

    # Find the row where "Total Parts Per Reports:" is in the DESCRIPTION column and update the QTY column with the total_qty
    df.loc[df['DESCRIPTION'].str.contains("Total Parts Per Reports:", na=False), 'QTY'] = total_qty
    df['BLANK'] = df['SEQ#']
    # Rename the 'BLANK' column to 'SEQ#' and the 'SEQ#' column to 'BLANK'
    df.rename(columns={'BLANK': 'TEMP', 'SEQ#': 'BLANK'}, inplace=True)
    df.rename(columns={'TEMP': 'SEQ#'}, inplace=True)

    # Make the original 'SEQ#' column (now named 'BLANK') empty
    df['BLANK'] = ''

    df['CODE'] = df['CODE'].apply(remove_number_after_second_asterisk)

    return df


def remove_number_after_second_asterisk(s):
    # Split the string by asterisks
    split_str = re.split(r'(\*)', s)

    # If there are at least three elements in the split string, remove anything after the second asterisk and join the string back together
    if len(split_str) >= 3:
        split_str = split_str[:3]
        s = ''.join(split_str)

    return s


def convert_pdf_to_excel(pdf_file, area_mm, column_positions_mm, output_file, modified_pdf_name):
    df = extract_tables_tabula(pdf_file, area_mm, column_positions_mm)
    df = clean_dataframe(df)
    df.to_excel(output_file, index=False)
    print("Converting")
    return df


def process_pdf(input_pdf, area_mm, column_positions_mm, output_excel):
    pdf_name = os.path.basename(input_pdf)[:-4]  # Get PDF file name without the file extension
    modified_pdf_name = pdf_name.replace(" - Part List", "")
    convert_pdf_to_excel(input_pdf, area_mm, column_positions_mm, output_excel, modified_pdf_name)
    format_excel_file(output_excel)  # Add this line to format the Excel file
    print("Processing")

# CORE FUNCTION V #
def process_pdf_directory(input_directory, area_mm, column_positions_mm):
    for filename in os.listdir(input_directory):
        if filename.endswith(".pdf"):
            input_pdf = os.path.join(input_directory, filename)
            output_excel = os.path.splitext(input_pdf)[0] + ".xlsx"
            process_pdf(input_pdf, area_mm, column_positions_mm, output_excel)



# if __name__ == '__main__':
#     # Create an argument parser
#     parser = argparse.ArgumentParser(description='Description of your program')
#     parser.add_argument('--input_directory', type=str, help='Path to the input directory')
#     parser.add_argument('--area_mm', type=float, help='Area of each box in mm^2')
#     parser.add_argument('--column_positions_mm', nargs='+', type=float, help='Positions of the columns in mm')
#     # Add more arguments as needed
#
#     # Parse the command-line arguments
#     args = parser.parse_args()

    # # Define the input_directory variable based on the command-line argument
    # input_directory = args.input_directory
    # print(input_directory)
    # # Use the command-line arguments in your script
    # area_mm = [0, 0, 196.85, 279.4]  # top, left, bottom, and right coordinates in millimeters
    # column_positions_mm = [15, 76, 90, 96, 98, 113, 164, 260]  # Approximate column positions in millimeters
    # # Use the values of the arguments in your script
    # process_pdf_directory(input_directory, area_mm, column_positions_mm)


# area_mm = [0, 0, 196.85, 279.4]  # top, left, bottom, and right coordinates in millimeters
# column_positions_mm = [15, 76, 90, 96, 98, 113, 164, 260]  # Approximate column positions in millimeters


#process_pdf_directory(input_directory, area_mm, column_positions_mm)